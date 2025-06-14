# rag_api/file_processing.py
import io
import logging
from pathlib import Path
import time  # Para IDs únicos

# Bibliotecas de extração
try:
    from PIL import Image
except ImportError:
    Image = None
    logging.warning("Pillow não instalado. OCR pode não funcionar.")
try:
    import pytesseract
except ImportError:
    pytesseract = None
    logging.warning("pytesseract não instalado. OCR de PDF não funcionará.")
try:
    from pdf2image import convert_from_bytes
except ImportError:
    convert_from_bytes = None
    logging.warning("pdf2image não instalado. OCR de PDF não funcionará.")
try:
    import pypdf
except ImportError:
    pypdf = None
    logging.warning("pypdf não instalado. Upload de PDF não funcionará.")
try:
    import docx
except ImportError:
    docx = None
    logging.warning("python-docx não instalado. Upload de DOCX não funcionará.")
try:
    import openpyxl
except ImportError:
    openpyxl = None
    logging.warning("openpyxl não instalado. Upload de XLSX não funcionará.")

logger = logging.getLogger(__name__)


def extract_text_from_pdf_with_ocr(file_content: bytes, lang: str = "por") -> str:
    """Tenta extrair texto, se falhar ou for vazio, usa OCR."""
    text = ""
    # Tenta extração direta primeiro (mais rápido se funcionar)
    try:
        if pypdf:
            reader = pypdf.PdfReader(io.BytesIO(file_content))
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
            text = text.strip()
            if text:  # Se extração direta funcionou, retorna
                logger.info("Texto extraído diretamente do PDF.")
                return text
        logger.warning(
            "Extração direta do PDF falhou ou retornou vazio. Tentando OCR..."
        )
    except Exception as e:
        logger.warning(f"Erro na extração direta do PDF: {e}. Prosseguindo com OCR.")
        text = ""  # Garante que text está vazio para OCR

    # Se chegou aqui, tenta OCR
    if not pytesseract or not convert_from_bytes or not Image:
        raise ImportError(
            "Bibliotecas necessárias para OCR (pytesseract, pdf2image, Pillow) não estão disponíveis."
        )

    ocr_text = ""
    try:
        images = convert_from_bytes(file_content, dpi=300)  # dpi pode ser ajustado
        for i, img in enumerate(images):
            logger.info(f"Processando OCR da página {i+1}...")
            page_ocr_text = pytesseract.image_to_string(img, lang=lang)
            if page_ocr_text:
                ocr_text += page_ocr_text + "\n"
        logger.info("OCR concluído.")
        return ocr_text.strip()
    except Exception as e:
        logger.error(f"Erro durante o processo de OCR: {e}", exc_info=True)
        # Retorna o texto da extração direta (mesmo que vazio) ou lança erro
        return text  # Ou raise ValueError(...)


def extract_text_from_docx(file_content: bytes) -> str:
    """Extrai texto de um conteúdo de arquivo DOCX em bytes."""
    if not docx:
        raise ImportError("Biblioteca python-docx é necessária para processar DOCX.")
    text = ""
    try:
        document = docx.Document(io.BytesIO(file_content))
        for para in document.paragraphs:
            text += para.text + "\n"
    except Exception as e:
        logger.error(f"Erro ao extrair texto do DOCX: {e}", exc_info=True)
        # raise ValueError(f"Não foi possível processar o DOCX: {e}") from e
    return text.strip()


def extract_text_from_xlsx(file_content: bytes) -> str:
    """Extrai texto de um conteúdo de arquivo XLSX em bytes."""
    if not openpyxl:
        raise ImportError("Biblioteca openpyxl é necessária para processar XLSX.")
    text = ""
    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(file_content), data_only=True
        )  # data_only=True pega valores, não fórmulas
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"--- Folha: {sheet_name} ---\n"
            for row in sheet.iter_rows():
                row_texts = [
                    str(cell.value) if cell.value is not None else "" for cell in row
                ]
                text += " | ".join(row_texts) + "\n"  # Separador simples entre células
            text += "\n"
    except Exception as e:
        logger.error(f"Erro ao extrair texto do XLSX: {e}", exc_info=True)
        # raise ValueError(f"Não foi possível processar o XLSX: {e}") from e
    return text.strip()


def extract_text_from_txt(file_content: bytes) -> str:
    """Extrai texto de um conteúdo de arquivo TXT em bytes, tentando UTF-8."""
    try:
        return file_content.decode("utf-8").strip()
    except UnicodeDecodeError:
        logger.warning("Falha ao decodificar TXT como UTF-8, tentando latin-1.")
        try:
            return file_content.decode("latin-1").strip()
        except Exception as e:
            logger.error(f"Erro ao decodificar TXT: {e}", exc_info=True)
            # raise ValueError(f"Não foi possível decodificar o TXT: {e}") from e
            return ""  # Ou retornar vazio


def extract_text_from_file(uploaded_file) -> tuple[str | None, str | None]:
    """
    Extrai texto de um arquivo carregado pelo Django.
    Retorna uma tupla (texto_extraido, erro_mensagem).
    Retorna (None, mensagem_erro) se a extração falhar ou tipo não suportado.
    """
    filename = uploaded_file.name
    file_content = uploaded_file.read()  # Ler o conteúdo em bytes
    file_extension = Path(filename).suffix.lower()

    logger.info(
        f"Processando arquivo '{filename}' com extensão '{file_extension}' ({len(file_content)} bytes)"
    )

    try:
        if file_extension == ".pdf":
            return extract_text_from_pdf_with_ocr(file_content), None
        elif file_extension == ".docx":
            return extract_text_from_docx(file_content), None
        elif file_extension == ".xlsx":
            return extract_text_from_xlsx(file_content), None
        elif file_extension == ".txt":
            return extract_text_from_txt(file_content), None
        # elif file_extension == ".odt":
        #     return extract_text_from_odt(file_content), None
        else:
            error_msg = f"Tipo de arquivo não suportado: {file_extension}"
            logger.warning(error_msg)
            return None, error_msg
    except ImportError as e:
        error_msg = f"Biblioteca necessária não encontrada para '{file_extension}': {e}"
        logger.error(error_msg)
        return None, error_msg
    except Exception as e:
        error_msg = f"Erro inesperado ao processar o arquivo '{filename}': {e}"
        logger.error(error_msg, exc_info=True)
        return None, error_msg


def simple_chunker(
    text: str, chunk_size: int = 1500, chunk_overlap: int = 150
) -> list[str]:
    """Divide o texto em pedaços (chunks) com sobreposição."""
    if not text:
        return []

    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end]
        chunks.append(chunk)
        start += chunk_size - chunk_overlap  # Move para o próximo chunk com overlap
        if start < 0:
            start = 0  # Evitar loop infinito com overlap grande
        if end >= len(text):  # Garante que o último caractere seja incluído
            break
    return chunks


def generate_chunk_ids(filename: str, num_chunks: int) -> list[str]:
    """Gera IDs únicos para os chunks de um arquivo."""
    # Usar timestamp para diferenciar uploads do mesmo arquivo
    timestamp = int(time.time())
    base_name = Path(filename).stem  # Nome do arquivo sem extensão
    # Simplificar/limitar tamanho do nome base para evitar IDs muito longos
    base_name_safe = "".join(c if c.isalnum() else "_" for c in base_name)[:50]
    return [f"{base_name_safe}_{timestamp}_chunk_{i}" for i in range(num_chunks)]
